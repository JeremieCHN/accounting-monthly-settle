"""月度汇算小工具 - Excel 读写逻辑"""

import logging
import os
from dataclasses import dataclass
from datetime import datetime
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


def col_letter(index: int) -> str:
    """将列索引(0-based)转为列字母，如 0->A, 1->B, 25->Z, 26->AA"""
    result = ""
    n = index
    while True:
        result = chr(n % 26 + ord("A")) + result
        n = n // 26 - 1
        if n < 0:
            break
    return result


from src.config import (
    ANOMALY_RULES,
    ANOMALY_SOURCE_APPLICABILITY,
    PREVIEW_DEFAULT_ROWS,
    InputSource,
    SOURCE_COLUMNS,
)

logger = logging.getLogger("monthly_settlement")


class ExcelHandler:
    """Excel 文件读取与解析"""

    def __init__(self, file_path: str):
        self.file_path = file_path
        self._workbook = None

    def _open(self):
        """延迟打开工作簿"""
        if self._workbook is None:
            try:
                self._workbook = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
            except Exception as e:
                logger.error(f"读取文件失败: {self.file_path}, 原因: {e}")
                raise FileNotFoundError(f"无法读取文件: {self.file_path}") from e

    def get_sheet_names(self) -> list[str]:
        """返回文件中所有 sheet 名称"""
        self._open()
        try:
            sheets = self._workbook.sheetnames
            logger.info(f"读取文件 {self.file_path} 的 sheet 列表: {sheets}")
            return sheets
        except Exception as e:
            logger.error(f"读取文件失败: {self.file_path}, 原因: {e}")
            raise

    def read_sheet(self, sheet_name: str, skip_header_rows: int = 0,
                    skip_footer_rows: int = 0) -> tuple[list[str], list[dict], int]:
        """
        读取指定 sheet 的数据
        skip_header_rows: 跳过开头的标题行数（在表头行之前）
        skip_footer_rows: 跳过末尾的汇总行数
        返回: (列名列表, 数据行列表, sheet原始总行数)
        每行数据为 {列名: 值} 的字典
        """
        self._open()
        try:
            ws = self._workbook[sheet_name]
        except KeyError:
            msg = f"Sheet '{sheet_name}' 不存在"
            logger.error(f"读取 sheet 失败: {sheet_name}, 原因: {msg}")
            raise ValueError(msg) from None

        rows_iter = ws.iter_rows(values_only=True)

        # 跳过开头的标题行
        for _ in range(skip_header_rows):
            try:
                next(rows_iter)
            except StopIteration:
                logger.info(f"读取 sheet: {sheet_name}, 行数: 0 (跳过标题行后无数据)")
                return [], [], skip_header_rows

        # 读取表头行
        try:
            header_row = next(rows_iter)
        except StopIteration:
            logger.info(f"读取 sheet: {sheet_name}, 行数: 0")
            return [], [], skip_header_rows

        columns = []
        for i, c in enumerate(header_row):
            if c is not None and str(c).strip():
                columns.append(str(c))
            else:
                columns.append(f"{col_letter(i)}列")

        # 读取所有数据行
        all_data = []
        for row in rows_iter:
            row_dict = {}
            for col_name, cell_value in zip(columns, row):
                row_dict[col_name] = cell_value
            all_data.append(row_dict)

        # 跳过末尾的汇总行
        if skip_footer_rows > 0:
            data = all_data[:-skip_footer_rows] if skip_footer_rows < len(all_data) else []
        else:
            data = all_data

        logger.info(f"读取 sheet: {sheet_name}, 行数: {len(data)}, 跳过标题行: {skip_header_rows}, 跳过汇总行: {skip_footer_rows}")
        total_rows = skip_header_rows + 1 + len(all_data)  # 标题行 + 表头行 + 所有数据行
        return columns, data, total_rows

    def close(self):
        """关闭工作簿"""
        if self._workbook is not None:
            self._workbook.close()
            self._workbook = None


class ColumnMapper:
    """列名自动匹配与手动映射"""

    def auto_match(self, sheet_columns: list[str], source: InputSource) -> dict[str, str | None]:
        """
        自动匹配: 将 sheet 列名映射到 source 的必需列
        返回: {column_key: matched_sheet_column_name | None}
        匹配规则: 精确匹配 aliases 中的任一名称
        """
        mapping: dict[str, str | None] = {}
        col_defs = SOURCE_COLUMNS[source]

        for col_def in col_defs:
            matched = None
            for alias in col_def.aliases:
                if alias in sheet_columns:
                    matched = alias
                    break
            mapping[col_def.key] = matched
            if matched is None:
                logger.warning(f"列 {col_def.key} 自动匹配失败，需手动选择")
            else:
                logger.debug(f"列 {col_def.key} 自动匹配到: {matched}")

        logger.debug(f"自动匹配结果: {mapping}")
        return mapping

    def validate_mapping(self, mapping: dict[str, str | None], source: InputSource) -> list[str]:
        """
        校验映射完整性
        返回: 未映射的必需列的 display 名称列表（空列表表示校验通过）
        """
        missing = []
        col_defs = SOURCE_COLUMNS[source]
        for col_def in col_defs:
            if col_def.required and (col_def.key not in mapping or mapping[col_def.key] is None):
                missing.append(col_def.display)
        return missing


@dataclass
class AnomalyRecord:
    """异常记录"""
    row_index: int       # 行号（从1开始，不含表头）
    column: str          # 列显示名
    value: Any           # 原始值
    rule: str            # 规则ID
    label: str           # 标注文本


class DataAnalyzer:
    """数据预览与异常检测"""

    def preview(self, data: list[dict], mapping: dict[str, str],
                rows: int = PREVIEW_DEFAULT_ROWS) -> list[dict]:
        """
        返回映射后的预览数据（仅包含已映射列）
        """
        preview_data = []
        for row in data[:rows]:
            mapped_row = {}
            for col_key, sheet_col in mapping.items():
                if sheet_col is not None:
                    mapped_row[col_key] = row.get(sheet_col)
            preview_data.append(mapped_row)
        return preview_data

    def get_stats(self, data: list[dict], mapping: dict[str, str]) -> dict:
        """
        返回统计信息: {total_rows, material_count}
        """
        total_rows = len(data)
        material_key = mapping.get("material")
        material_count = 0
        if material_key:
            materials = set()
            for row in data:
                val = row.get(material_key)
                if val is not None:
                    materials.add(str(val))
            material_count = len(materials)
        return {"total_rows": total_rows, "material_count": material_count}

    def detect_anomalies(self, data: list[dict], mapping: dict[str, str],
                         source: InputSource, row_offset: int = 1) -> list[AnomalyRecord]:
        """
        检测异常数据
        row_offset: 行号偏移量（跳过标题行数 + 1表头行），用于计算 Excel 行号
        返回异常记录列表，每条包含: {row_index, column, value, rule, label}
        """
        anomalies: list[AnomalyRecord] = []

        # 获取适用于当前来源的规则
        applicable_rules = {
            rule_id: rule for rule_id, rule in ANOMALY_RULES.items()
            if source in ANOMALY_SOURCE_APPLICABILITY.get(rule_id, [])
        }

        # 构建 key -> ColumnDef 的映射
        col_def_map = {col_def.key: col_def for col_def in SOURCE_COLUMNS[source]}

        for row_idx, row in enumerate(data):
            excel_row = row_offset + row_idx + 1  # +1 因为数据行从1开始
            for rule_id, rule in applicable_rules.items():
                col_key = rule["key"]
                sheet_col = mapping.get(col_key)
                if sheet_col is None:
                    continue
                value = row.get(sheet_col)
                condition = rule["condition"]

                is_anomaly = False
                if condition == "<0":
                    if isinstance(value, (int, float)) and value < 0:
                        is_anomaly = True
                elif condition == "==0":
                    if isinstance(value, (int, float)) and value == 0:
                        is_anomaly = True
                elif condition == "is_empty":
                    if value is None or (isinstance(value, str) and value.strip() == ""):
                        is_anomaly = True

                if is_anomaly:
                    col_def = col_def_map.get(col_key)
                    display_name = col_def.display if col_def else col_key
                    anomalies.append(AnomalyRecord(
                        row_index=excel_row,
                        column=display_name,
                        value=value,
                        rule=rule_id,
                        label=rule["label"],
                    ))

        if anomalies:
            logger.warning(f"检测到 {len(anomalies)} 条异常数据")
        return anomalies


class ExcelExporter:
    """Excel 结果导出"""

    def export(self, result, output_path: str) -> None:
        """
        将计算结果导出为 xlsx 文件
        输出文件包含2个 sheet:
          - 期末库存: 物料名称、数量、单价、金额
          - 出库成本: 物料名称、出库数量、单价、出库金额
        """
        from src.fifo_engine import CalcResult

        try:
            # 确保输出目录存在
            output_dir = os.path.dirname(output_path)
            if output_dir:
                os.makedirs(output_dir, exist_ok=True)

            wb = openpyxl.Workbook()

            # 样式定义
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            warning_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            left_align = Alignment(horizontal="left")
            right_align = Alignment(horizontal="right")
            center_align = Alignment(horizontal="center")

            # ---- Sheet 1: 期末库存 ----
            ws_closing = wb.active
            ws_closing.title = "期末库存"

            closing_headers = ["物料名称", "数量", "单价", "金额"]
            closing_widths = [20, 12, 12, 14]

            for col_idx, (header, width) in enumerate(zip(closing_headers, closing_widths), 1):
                cell = ws_closing.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                ws_closing.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

            total_closing_qty = 0.0
            total_closing_amount = 0.0

            for row_idx, m in enumerate(result.materials, 2):
                has_warning = len(m.warnings) > 0
                fill = warning_fill if has_warning else None

                cell_name = ws_closing.cell(row=row_idx, column=1, value=m.material_name)
                cell_name.alignment = left_align
                if fill:
                    cell_name.fill = fill

                cell_qty = ws_closing.cell(row=row_idx, column=2, value=m.closing_quantity)
                cell_qty.alignment = right_align
                cell_qty.number_format = "0.00"
                if fill:
                    cell_qty.fill = fill

                cell_price = ws_closing.cell(row=row_idx, column=3, value=m.closing_avg_price)
                cell_price.alignment = right_align
                cell_price.number_format = "0.00"
                if fill:
                    cell_price.fill = fill

                cell_amount = ws_closing.cell(row=row_idx, column=4, value=m.closing_amount)
                cell_amount.alignment = right_align
                cell_amount.number_format = "0.00"
                if fill:
                    cell_amount.fill = fill

                total_closing_qty += m.closing_quantity
                total_closing_amount += m.closing_amount

            # 汇总行
            summary_row = len(result.materials) + 2
            ws_closing.cell(row=summary_row, column=1, value="合计").font = header_font
            cell_sum_qty = ws_closing.cell(row=summary_row, column=2, value=round(total_closing_qty, 2))
            cell_sum_qty.font = header_font
            cell_sum_qty.alignment = right_align
            cell_sum_qty.number_format = "0.00"
            ws_closing.cell(row=summary_row, column=3).font = header_font
            cell_sum_amount = ws_closing.cell(row=summary_row, column=4, value=round(total_closing_amount, 2))
            cell_sum_amount.font = header_font
            cell_sum_amount.alignment = right_align
            cell_sum_amount.number_format = "0.00"

            # ---- Sheet 2: 出库成本 ----
            ws_outbound = wb.create_sheet("出库成本")

            outbound_headers = ["物料名称", "出库数量", "单价", "出库金额"]
            outbound_widths = [20, 12, 12, 14]

            for col_idx, (header, width) in enumerate(zip(outbound_headers, outbound_widths), 1):
                cell = ws_outbound.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                ws_outbound.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

            total_outbound_qty = 0.0
            total_outbound_amount = 0.0

            for row_idx, m in enumerate(result.materials, 2):
                has_warning = len(m.warnings) > 0
                fill = warning_fill if has_warning else None

                cell_name = ws_outbound.cell(row=row_idx, column=1, value=m.material_name)
                cell_name.alignment = left_align
                if fill:
                    cell_name.fill = fill

                cell_qty = ws_outbound.cell(row=row_idx, column=2, value=m.outbound_quantity)
                cell_qty.alignment = right_align
                cell_qty.number_format = "0.00"
                if fill:
                    cell_qty.fill = fill

                cell_price = ws_outbound.cell(row=row_idx, column=3, value=m.outbound_avg_price)
                cell_price.alignment = right_align
                cell_price.number_format = "0.00"
                if fill:
                    cell_price.fill = fill

                cell_amount = ws_outbound.cell(row=row_idx, column=4, value=m.outbound_amount)
                cell_amount.alignment = right_align
                cell_amount.number_format = "0.00"
                if fill:
                    cell_amount.fill = fill

                total_outbound_qty += m.outbound_quantity
                total_outbound_amount += m.outbound_amount

            # 汇总行
            summary_row = len(result.materials) + 2
            ws_outbound.cell(row=summary_row, column=1, value="合计").font = header_font
            cell_sum_qty = ws_outbound.cell(row=summary_row, column=2, value=round(total_outbound_qty, 2))
            cell_sum_qty.font = header_font
            cell_sum_qty.alignment = right_align
            cell_sum_qty.number_format = "0.00"
            ws_outbound.cell(row=summary_row, column=3).font = header_font
            cell_sum_amount = ws_outbound.cell(row=summary_row, column=4, value=round(total_outbound_amount, 2))
            cell_sum_amount.font = header_font
            cell_sum_amount.alignment = right_align
            cell_sum_amount.number_format = "0.00"

            wb.save(output_path)
            logger.info(f"导出结果到文件: {output_path}")

        except Exception as e:
            logger.error(f"导出文件失败: {output_path}, 原因: {e}")
            raise
