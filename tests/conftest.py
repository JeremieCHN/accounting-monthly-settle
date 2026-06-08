"""公共 fixtures"""

import pytest
import openpyxl

from src.config import InputSource
from src.fifo_engine import FIFOEngine


@pytest.fixture
def sample_xlsx(tmp_path):
    """创建一个包含测试数据的 xlsx 文件，返回文件路径"""
    file_path = tmp_path / "test_data.xlsx"
    wb = openpyxl.Workbook()

    # 入库记录 sheet
    ws_inbound = wb.active
    ws_inbound.title = "入库记录"
    ws_inbound.append(["进货日期", "物料名称", "数量", "单价"])
    ws_inbound.append(["2026-01-01", "物料A", 100, 10.0])
    ws_inbound.append(["2026-01-02", "物料B", 50, 20.0])
    ws_inbound.append(["2026-01-03", "物料A", -5, 10.0])   # 数量为负
    ws_inbound.append(["", "物料C", 30, 0])                  # 日期为空，单价为0
    ws_inbound.append(["2026-01-05", "物料B", 20, 15.0])

    # 期初库存 sheet
    ws_opening = wb.create_sheet("期初库存")
    ws_opening.append(["物料名称", "数量", "单价"])
    ws_opening.append(["物料A", 200, 8.0])
    ws_opening.append(["物料B", 100, 18.0])
    ws_opening.append(["物料C", 50, 0])     # 单价为0
    ws_opening.append(["物料D", -10, 5.0])  # 数量为负

    # 出库记录 sheet
    ws_outbound = wb.create_sheet("出库记录")
    ws_outbound.append(["物料名称", "数量"])
    ws_outbound.append(["物料A", 80])
    ws_outbound.append(["物料B", 30])
    ws_outbound.append(["物料C", -5])  # 数量为负
    ws_outbound.append(["物料A", 50])

    wb.save(str(file_path))
    return str(file_path)


@pytest.fixture
def inbound_columns():
    """返回入库记录的典型列名列表"""
    return ["进货日期", "物料名称", "数量", "单价"]


@pytest.fixture
def sample_mapping():
    """返回一个完整的列映射字典"""
    return {
        "date": "进货日期",
        "material": "物料名称",
        "quantity": "数量",
        "price": "单价",
    }


# ---- FIFO 测试 fixtures (里程碑二) ----

@pytest.fixture
def fifo_engine():
    """FIFO 计算引擎实例"""
    return FIFOEngine()


@pytest.fixture
def opening_data():
    """期初库存映射后数据"""
    return [
        {"material": "蒸笼",   "quantity": 23, "price": 7.1},
        {"material": "蒸笼盖", "quantity": 2,  "price": 7.1},
    ]


@pytest.fixture
def inbound_data():
    """入库记录映射后数据"""
    return [
        {"date": "2026/4/1", "material": "蒸笼",   "quantity": 23,  "price": 7.1},
        {"date": "2026/4/1", "material": "蒸笼盖", "quantity": 2,   "price": 7.1},
        {"date": "2026/4/1", "material": "日晒盐", "quantity": 200, "price": 61.5},
    ]


@pytest.fixture
def outbound_data():
    """出库记录映射后数据"""
    return [
        {"material": "蒸笼",   "quantity": 23},
        {"material": "蒸笼盖", "quantity": 2},
    ]
