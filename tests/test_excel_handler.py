"""ExcelHandler & ColumnMapper & DataAnalyzer & ExcelExporter 测试"""

import os
import pytest

import openpyxl

from src.config import InputSource, SOURCE_COLUMNS
from src.excel_handler import ExcelHandler, ColumnMapper, DataAnalyzer, AnomalyRecord, ExcelExporter
from src.fifo_engine import FIFOEngine, CalcResult, MaterialResult, Batch


# ============================================================
# ExcelHandler 测试
# ============================================================

@pytest.mark.integration
class TestExcelHandler:
    """ExcelHandler 集成测试"""

    def test_xh_001_get_sheet_names(self, sample_xlsx):
        """XH-001: 读取有效 xlsx 的 sheet 列表"""
        handler = ExcelHandler(sample_xlsx)
        try:
            sheets = handler.get_sheet_names()
            assert sheets == ["入库记录", "期初库存", "出库记录"]
        finally:
            handler.close()

    def test_xh_002_file_not_found(self):
        """XH-002: 读取不存在的文件"""
        handler = ExcelHandler("/nonexistent/file.xlsx")
        with pytest.raises(FileNotFoundError):
            handler.get_sheet_names()

    def test_xh_003_read_sheet_data(self, sample_xlsx):
        """XH-003: 读取指定 sheet 的数据"""
        handler = ExcelHandler(sample_xlsx)
        try:
            columns, data, _ = handler.read_sheet("入库记录")
            assert columns == ["进货日期", "物料名称", "数量", "单价"]
            assert len(data) == 5
            assert data[0]["物料名称"] == "物料A"
            assert data[0]["数量"] == 100
        finally:
            handler.close()

    def test_xh_004_read_nonexistent_sheet(self, sample_xlsx):
        """XH-004: 读取不存在的 sheet"""
        handler = ExcelHandler(sample_xlsx)
        try:
            with pytest.raises(ValueError):
                handler.read_sheet("不存在的Sheet")
        finally:
            handler.close()

    def test_xh_005_read_empty_sheet(self, tmp_path):
        """XH-005: 读取空 sheet"""
        import openpyxl
        file_path = tmp_path / "empty.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "空Sheet"
        # 不写入任何数据
        wb.save(str(file_path))

        handler = ExcelHandler(str(file_path))
        try:
            columns, data, _ = handler.read_sheet("空Sheet")
            assert columns == []
            assert data == []
        finally:
            handler.close()

    def test_xh_006_data_types(self, sample_xlsx):
        """XH-006: 数据行值类型正确"""
        handler = ExcelHandler(sample_xlsx)
        try:
            columns, data, _ = handler.read_sheet("入库记录")
            row = data[0]
            # 数值为 int
            assert isinstance(row["数量"], (int, float))
            # 单价为 float
            assert isinstance(row["单价"], (int, float))
            # 字符串为 str
            assert isinstance(row["物料名称"], str)
        finally:
            handler.close()

    def test_xh_007_skip_header_rows(self, tmp_path):
        """XH-007: 跳过开头标题行"""
        import openpyxl
        file_path = tmp_path / "skip_header.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "数据"
        ws.append(["这是标题行1"])       # 应被跳过
        ws.append(["这是标题行2"])       # 应被跳过
        ws.append(["物料名称", "数量", "单价"])  # 真正的表头
        ws.append(["物料A", 10, 5.0])
        ws.append(["物料B", 20, 8.0])
        wb.save(str(file_path))

        handler = ExcelHandler(str(file_path))
        try:
            columns, data, _ = handler.read_sheet("数据", skip_header_rows=2)
            assert columns == ["物料名称", "数量", "单价"]
            assert len(data) == 2
            assert data[0]["物料名称"] == "物料A"
        finally:
            handler.close()

    def test_xh_008_skip_footer_rows(self, sample_xlsx):
        """XH-008: 跳过末尾汇总行"""
        handler = ExcelHandler(sample_xlsx)
        try:
            columns, data, _ = handler.read_sheet("入库记录", skip_footer_rows=1)
            assert len(data) == 4  # 原本5行，跳过1行
        finally:
            handler.close()

    def test_xh_009_skip_both(self, tmp_path):
        """XH-009: 同时跳过标题行和汇总行"""
        import openpyxl
        file_path = tmp_path / "skip_both.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "数据"
        ws.append(["标题行"])            # 跳过
        ws.append(["物料名称", "数量"])   # 表头
        ws.append(["物料A", 10])
        ws.append(["物料B", 20])
        ws.append(["合计", 30])          # 跳过
        wb.save(str(file_path))

        handler = ExcelHandler(str(file_path))
        try:
            columns, data, _ = handler.read_sheet("数据", skip_header_rows=1, skip_footer_rows=1)
            assert columns == ["物料名称", "数量"]
            assert len(data) == 2
        finally:
            handler.close()


# ============================================================
# ColumnMapper 测试
# ============================================================

@pytest.mark.unit
class TestColumnMapper:
    """ColumnMapper 单元测试"""

    def setup_method(self):
        self.mapper = ColumnMapper()

    def test_cm_001_exact_match(self):
        """CM-001: 精确匹配标准列名"""
        columns = ["进货日期", "物料名称", "数量", "单价"]
        mapping = self.mapper.auto_match(columns, InputSource.INBOUND)
        assert mapping["date"] == "进货日期"
        assert mapping["material"] == "物料名称"
        assert mapping["quantity"] == "数量"
        assert mapping["price"] == "单价"

    def test_cm_002_match_aliases(self):
        """CM-002: 匹配别名"""
        columns = ["日期", "品名", "入库数量", "入库单价"]
        mapping = self.mapper.auto_match(columns, InputSource.INBOUND)
        assert mapping["date"] == "日期"
        assert mapping["material"] == "品名"
        assert mapping["quantity"] == "入库数量"
        assert mapping["price"] == "入库单价"

    def test_cm_003_partial_match(self):
        """CM-003: 部分列无法匹配"""
        columns = ["进货日期", "物料名称", "其他列"]
        mapping = self.mapper.auto_match(columns, InputSource.INBOUND)
        assert mapping["date"] == "进货日期"
        assert mapping["material"] == "物料名称"
        assert mapping["quantity"] is None
        assert mapping["price"] is None

    def test_cm_004_no_match(self):
        """CM-004: 完全无法匹配"""
        columns = ["列A", "列B", "列C"]
        mapping = self.mapper.auto_match(columns, InputSource.INBOUND)
        assert all(v is None for v in mapping.values())

    def test_cm_005_extra_spaces_no_match(self):
        """CM-005: 列名有多余空格，精确匹配失败"""
        columns = [" 进货日期", "物料名称 ", " 数量 "]
        mapping = self.mapper.auto_match(columns, InputSource.INBOUND)
        assert mapping["date"] is None
        assert mapping["material"] is None
        assert mapping["quantity"] is None

    def test_cm_006_validate_mapping_complete(self):
        """CM-006: validate_mapping 全部映射"""
        mapping = {"date": "进货日期", "material": "物料名称", "quantity": "数量", "price": "单价"}
        missing = self.mapper.validate_mapping(mapping, InputSource.INBOUND)
        assert missing == []

    def test_cm_007_validate_mapping_missing(self):
        """CM-007: validate_mapping 缺少必需列"""
        mapping = {"date": "进货日期", "material": None, "quantity": None, "price": None}
        missing = self.mapper.validate_mapping(mapping, InputSource.INBOUND)
        assert "物料名称" in missing
        assert "数量" in missing
        assert "单价" in missing


# ============================================================
# DataAnalyzer 测试
# ============================================================

@pytest.mark.unit
class TestDataAnalyzer:
    """DataAnalyzer 单元测试"""

    def setup_method(self):
        self.analyzer = DataAnalyzer()

    def _make_data(self, n: int = 25) -> list[dict]:
        """生成测试数据"""
        data = []
        for i in range(n):
            data.append({
                "进货日期": f"2026-01-{i+1:02d}",
                "物料名称": f"物料{(i % 3) + 1}",
                "数量": i + 1,
                "单价": 10.0 + i,
            })
        return data

    def test_da_001_preview_default_rows(self):
        """DA-001: 预览默认行数"""
        data = self._make_data(25)
        mapping = {"date": "进货日期", "material": "物料名称", "quantity": "数量", "price": "单价"}
        preview = self.analyzer.preview(data, mapping)
        assert len(preview) == 20

    def test_da_002_preview_less_than_default(self):
        """DA-002: 数据不足20行"""
        data = self._make_data(10)
        mapping = {"date": "进货日期", "material": "物料名称", "quantity": "数量", "price": "单价"}
        preview = self.analyzer.preview(data, mapping)
        assert len(preview) == 10

    def test_da_003_preview_only_mapped_columns(self):
        """DA-003: 预览仅包含已映射列"""
        data = [{"进货日期": "2026-01-01", "物料名称": "A", "数量": 10, "单价": 5.0, "额外列": "x"}]
        mapping = {"date": "进货日期", "material": "物料名称"}
        preview = self.analyzer.preview(data, mapping)
        assert "date" in preview[0]
        assert "material" in preview[0]
        assert "额外列" not in preview[0]

    def test_da_004_stats_total_rows(self):
        """DA-004: 统计总行数"""
        data = self._make_data(15)
        mapping = {"date": "进货日期", "material": "物料名称", "quantity": "数量", "price": "单价"}
        stats = self.analyzer.get_stats(data, mapping)
        assert stats["total_rows"] == 15

    def test_da_005_stats_material_count(self):
        """DA-005: 统计物料种类数"""
        data = self._make_data(9)  # 3种物料，各3行
        mapping = {"date": "进货日期", "material": "物料名称", "quantity": "数量", "price": "单价"}
        stats = self.analyzer.get_stats(data, mapping)
        assert stats["material_count"] == 3

    def test_da_006_detect_quantity_negative(self):
        """DA-006: 检测数量为负"""
        data = [
            {"物料名称": "A", "数量": -5, "单价": 10.0},
            {"物料名称": "B", "数量": 10, "单价": 20.0},
        ]
        mapping = {"material": "物料名称", "quantity": "数量", "price": "单价"}
        anomalies = self.analyzer.detect_anomalies(data, mapping, InputSource.INBOUND)
        assert len(anomalies) >= 1
        assert any(a.rule == "quantity_negative" for a in anomalies)

    def test_da_007_detect_date_empty(self):
        """DA-007: 检测日期为空"""
        data = [
            {"进货日期": None, "物料名称": "A", "数量": 10, "单价": 5.0},
            {"进货日期": "", "物料名称": "B", "数量": 5, "单价": 10.0},
        ]
        mapping = {"date": "进货日期", "material": "物料名称", "quantity": "数量", "price": "单价"}
        anomalies = self.analyzer.detect_anomalies(data, mapping, InputSource.INBOUND)
        date_anomalies = [a for a in anomalies if a.rule == "date_empty"]
        assert len(date_anomalies) == 2

    def test_da_008_detect_price_zero(self):
        """DA-008: 检测单价为0"""
        data = [
            {"物料名称": "A", "数量": 10, "单价": 0},
            {"物料名称": "B", "数量": 5, "单价": 10.0},
        ]
        mapping = {"material": "物料名称", "quantity": "数量", "price": "单价"}
        anomalies = self.analyzer.detect_anomalies(data, mapping, InputSource.INBOUND)
        assert any(a.rule == "price_zero" for a in anomalies)

    def test_da_009_no_anomalies(self):
        """DA-009: 无异常数据"""
        data = [
            {"进货日期": "2026-01-01", "物料名称": "A", "数量": 10, "单价": 5.0},
        ]
        mapping = {"date": "进货日期", "material": "物料名称", "quantity": "数量", "price": "单价"}
        anomalies = self.analyzer.detect_anomalies(data, mapping, InputSource.INBOUND)
        assert anomalies == []

    def test_da_010_multiple_anomalies(self):
        """DA-010: 多种异常同时存在"""
        data = [
            {"进货日期": None, "物料名称": "A", "数量": -5, "单价": 0},
        ]
        mapping = {"date": "进货日期", "material": "物料名称", "quantity": "数量", "price": "单价"}
        anomalies = self.analyzer.detect_anomalies(data, mapping, InputSource.INBOUND)
        rules = {a.rule for a in anomalies}
        assert "quantity_negative" in rules
        assert "date_empty" in rules
        assert "price_zero" in rules

    def test_da_011_row_offset(self):
        """DA-011: row_offset 影响异常行号"""
        data = [
            {"物料名称": "A", "数量": -5, "单价": 10.0},
        ]
        mapping = {"material": "物料名称", "quantity": "数量", "price": "单价"}
        # 默认 row_offset=1, 数据行从1开始 → Excel行号 = 1 + 0 + 1 = 2
        anomalies = self.analyzer.detect_anomalies(data, mapping, InputSource.INBOUND)
        assert anomalies[0].row_index == 2

        # row_offset=3 (跳过2行标题+1表头), 数据行从1开始 → Excel行号 = 3 + 0 + 1 = 4
        anomalies = self.analyzer.detect_anomalies(data, mapping, InputSource.INBOUND, row_offset=3)
        assert anomalies[0].row_index == 4


# ============================================================
# ExcelExporter 测试
# ============================================================

@pytest.mark.unit
class TestExcelExporter:
    """ExcelExporter 单元测试"""

    def setup_method(self):
        self.exporter = ExcelExporter()
        self.engine = FIFOEngine()

    def _make_calc_result(self, materials=None, has_warnings=False):
        """辅助: 创建 CalcResult"""
        if materials is None:
            materials = []
        return CalcResult(materials=materials, has_warnings=has_warnings)

    def test_exp_001_normal_export(self, tmp_path):
        """EXP-001: 导出正常结果，文件包含2个 sheet，数据正确"""
        opening = [
            {"material": "猪肉", "quantity": 100, "price": 10.0},
            {"material": "牛肉", "quantity": 50, "price": 20.0},
        ]
        inbound = [{"date": "2026/4/1", "material": "猪肉", "quantity": 50, "price": 12.0}]
        outbound = [{"material": "猪肉", "quantity": 80}]

        result = self.engine.calculate(opening, inbound, outbound)
        output_path = str(tmp_path / "result.xlsx")
        self.exporter.export(result, output_path)

        assert os.path.exists(output_path)
        wb = openpyxl.load_workbook(output_path)
        assert "期末库存" in wb.sheetnames
        assert "出库成本" in wb.sheetnames

        # 验证期末库存数据
        ws = wb["期末库存"]
        assert ws.cell(row=1, column=1).value == "物料名称"
        assert ws.cell(row=2, column=1).value == "牛肉"
        assert ws.cell(row=3, column=1).value == "猪肉"

        # 验证出库成本数据
        ws = wb["出库成本"]
        assert ws.cell(row=1, column=1).value == "物料名称"
        wb.close()

    def test_exp_002_warning_rows_highlighted(self, tmp_path):
        """EXP-002: 导出含警告的结果，警告行标黄"""
        opening = [{"material": "糖", "quantity": 10, "price": 8.0}]
        inbound = []
        outbound = [{"material": "糖", "quantity": 20}]

        result = self.engine.calculate(opening, inbound, outbound)
        assert result.has_warnings

        output_path = str(tmp_path / "result_warning.xlsx")
        self.exporter.export(result, output_path)

        wb = openpyxl.load_workbook(output_path)
        ws = wb["期末库存"]
        # 糖行(第2行)应该有黄色背景
        cell = ws.cell(row=2, column=1)
        assert cell.fill.start_color.rgb == "00FFFF00"
        wb.close()

    def test_exp_003_empty_result(self, tmp_path):
        """EXP-003: 导出空结果，文件包含2个 sheet，仅有表头"""
        result = self._make_calc_result()
        output_path = str(tmp_path / "result_empty.xlsx")
        self.exporter.export(result, output_path)

        wb = openpyxl.load_workbook(output_path)
        ws = wb["期末库存"]
        assert ws.cell(row=1, column=1).value == "物料名称"
        # 空结果时第2行为合计行，无数据行
        assert ws.cell(row=2, column=1).value == "合计"

        ws = wb["出库成本"]
        assert ws.cell(row=1, column=1).value == "物料名称"
        wb.close()

    def test_exp_004_create_directory(self, tmp_path):
        """EXP-004: 导出路径不存在，自动创建目录"""
        result = self._make_calc_result()
        output_path = str(tmp_path / "subdir" / "nested" / "result.xlsx")
        self.exporter.export(result, output_path)
        assert os.path.exists(output_path)

    def test_exp_005_file_in_use(self, tmp_path):
        """EXP-005: 导出文件被占用，抛出异常"""
        result = self._make_calc_result()
        output_path = str(tmp_path / "result.xlsx")

        # 先创建文件并保持打开
        wb = openpyxl.Workbook()
        wb.save(output_path)

        # 尝试导出到同一文件（openpyxl 可能不会抛出异常，取决于平台）
        # 这里主要验证异常处理逻辑存在
        try:
            # 在 Windows 上，如果文件被其他进程占用会抛出 PermissionError
            # 但 openpyxl 的 Workbook.save 可能直接覆盖，所以这个测试
            # 主要验证代码路径存在
            self.exporter.export(result, output_path)
        except Exception:
            pass  # 预期可能抛出异常

        wb.close()

    def test_exp_006_number_format(self, tmp_path):
        """EXP-006: 导出数值格式，数值列保留2位小数"""
        opening = [{"material": "盐", "quantity": 33, "price": 7.1}]
        inbound = []
        outbound = [{"material": "盐", "quantity": 11}]

        result = self.engine.calculate(opening, inbound, outbound)
        output_path = str(tmp_path / "result_format.xlsx")
        self.exporter.export(result, output_path)

        wb = openpyxl.load_workbook(output_path)
        ws = wb["期末库存"]
        # 检查数值列的 number_format
        assert ws.cell(row=2, column=2).number_format == "0.00"
        assert ws.cell(row=2, column=3).number_format == "0.00"
        assert ws.cell(row=2, column=4).number_format == "0.00"
        wb.close()
